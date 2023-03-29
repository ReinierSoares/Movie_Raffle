import PySimpleGUI as sg
import pandas as pd
from random import randint
import webbrowser
import openpyxl
import os


def random_film(genero, check):
    global window
    global link
    global movie

    while True:
        random_number = randint(0, len(movie_list)+1)

        if check:
            if genero == 'Todos':
                break
            else:
                if genero == gender_list[random_number]:
                    break
        else:
            check_movie = check_list[random_number].lower()
            if genero == 'Todos' and check_movie != "sim":
                break
            else:
                if genero == gender_list[random_number] and check_movie != "sim":
                    break

    movie = movie_list[random_number]
    link = link_list[random_number]
    window["-TEXTO-"].update(movie)


def add_film_window():
    global window
    items = ['Romance', 'Terror', 'Heróis', 'Animação',
             'Ação/Aventura', 'Suspense']
    layout = [[sg.Menu(menu_layout)],
              [sg.Text('Selecione um gênero:')],
              [sg.Combo(values=items, key='-GENDER-', size=(20, 1))],
              [sg.Text('Nome do Filme:'), sg.InputText(key='-FILM-')],
              [sg.Text('Link do Filme:'), sg.InputText(key='-LINK-')],
              [sg.Button('Ok'), sg.Button('Cancelar')]]

    window = sg.Window('Adicionar Filme', layout, use_ttk_buttons=True)

    while True:
        event, values = window.read()
        try:
            if event == sg.WIN_CLOSED or event == 'Sair':
                break
            elif event == 'Cancelar':
                window.close()
                main_window()
            elif event == 'Ok':
                gender_value = values['-GENDER-']
                film_value = values['-FILM-']
                link_value = values['-LINK-']

                if film_value == '':
                    pass
                else:
                    add_movie(gender_value, film_value, link_value)
            menu_bar(event, 'add')
        except:
            pass


def open_film_window(movie, link, check, row):
    global table

    if check.lower() == '✔️':
        check_button = sg.Button('Desmarcar visto')
    else:
        check_button = sg.Button('Marcar visto')

    layout = [[sg.Text(movie)],
              [sg.Button('Abrir Link'), check_button, sg.Button('Cancelar')]]

    window = sg.Window('Filme', layout, use_ttk_buttons=True)
    # Mostra a janela e aguarda a resposta do usuário
    button, _ = window.Read()

    # Verifica qual botão foi pressionado
    if button == 'Abrir Link':
        open_link(link, movie)

    elif button == 'Marcar visto':
        workbook = openpyxl.load_workbook(file_path)

        worksheet = workbook.active

        worksheet[f'D{row+2}'] = "Sim"

        workbook.save(file_path)
        att_movies()
    elif button == 'Desmarcar visto':
        workbook = openpyxl.load_workbook(file_path)

        worksheet = workbook.active

        worksheet[f'D{row+2}'] = "Não"

        workbook.save(file_path)
        att_movies()
    else:
        pass
    df = open_excel()
    column_to_remove = df['Link']
    df = df.drop(columns=['Link'])

    data_list = df.values.tolist()
    window.Close()
    return data_list


def open_excel():
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        sg.popup(e)
    return df


def data_check_list(data_list):
    for linha in data_list:
        if linha[2].lower() == 'sim':  # verifique o valor da segunda coluna
            linha.insert(2, '✔️')
        else:
            linha.insert(2, '☐')
    return data_list


def list_window():
    global window

    df = open_excel()
    column_to_remove = df['Link']
    df = df.drop(columns=['Link'])

    data_list = df.values.tolist()
    data_list = data_check_list(data_list)

    layout = [
        [sg.Menu(menu_layout)],
        [sg.InputText(key='-SEARCH-'), sg.Button('Pesquisar'),
         sg.Button('Limpar')],
        [sg.Table(
            values=data_list,
            headings=df.columns.tolist(),
            max_col_width=70,
            auto_size_columns=True,
            justification='left',
            num_rows=min(len(data_list), 20),
            row_height=30,
            key='-TABLE-',
            enable_events=True
        )]
    ]

    window = sg.Window("Tabela de Filmes", auto_size_text=True, auto_size_buttons=True,
                       grab_anywhere=False, resizable=True,
                       layout=layout, finalize=True, use_ttk_buttons=True)
    window['-TABLE-'].expand(True, True)
    window['-TABLE-'].table_frame.pack(expand=True, fill='both')
    search = False
    while True:
        event, values = window.read()

        try:
            if event == 'Sair' or event == sg.WIN_CLOSED:
                break
            if event == '-TABLE-':

                if values['-TABLE-']:
                    if search is True:
                        selected_row = index
                    else:
                        selected_row = values['-TABLE-'][0]
                    table_movie = data_list[selected_row][1]
                    table_link = column_to_remove[selected_row]
                    table_check = data_list[selected_row][2]
                    data_list = open_film_window(
                        table_movie, table_link, table_check, selected_row)
                    data_list = data_check_list(data_list)
                    window['-TABLE-'].update(values=data_list)
                    search = False
            elif event == 'Pesquisar':

                search_term = values['-SEARCH-']
                if search_term:
                    index, filtered_data = search_in_table(search_term)
                    data_search = filtered_data.values.tolist()
                    data_search = data_check_list(data_search)
                    window['-TABLE-'].update(values=data_search)
                    search = True
                else:
                    search = False
                    window['-TABLE-'].update(values=data_list)
            elif event == 'Limpar':
                search = False
                window['-TABLE-'].update(values=data_list)
                window['-SEARCH-'].update("")
            menu_bar(event, 'list')
        except:
            pass
    window.close()


def search_in_table(search_term):
    df = open_excel()
    column_to_remove = df['Link']
    df = df.drop(columns=['Link'])
    filtered_data = df[df.apply(lambda x: search_term.lower(
    ) in x.astype(str).str.lower().values.tolist(), axis=1)]
    if not filtered_data.empty:
        index = filtered_data.index[0]
        return index, filtered_data
    else:
        return None, None


def main_window():
    global window
    global link
    global movie

    # Define the dictionary of items
    items = ['Todos', 'Romance', 'Terror', 'Heróis', 'Animação',
             'Ação/Aventura', 'Suspense']
    link = None
    movie = None
    # Define the layout of your window
    layout = [[sg.Menu(menu_layout)],
              [sg.Text('Selecione um gênero:')],
              [sg.Combo(values=items, key='-GENDER-', size=(20, 1)),
               sg.Button('Sortear'), sg.Checkbox(
                  'Deseja ver filmes já vistos?', key='-CHECK-', default=True)],
              [sg.Text("Filme:"), sg.Text(key='-TEXTO-')],
              [sg.Button('Open link')]]

    # Create the window
    window = sg.Window('Sorteador de Filmes', layout,
                       finalize=True, use_ttk_buttons=True)

    while True:
        event, values = window.read()

        try:
            if event == 'Sortear':
                selected_item = values['-GENDER-']
                check_movie = values['-CHECK-']
                if selected_item != '':
                    random_film(selected_item, check_movie)
                    window.refresh()

            elif event == 'Open link':
                if link is None:
                    pass
                else:
                    open_link(link, movie)
            elif event == sg.WIN_CLOSED or event == 'Sair':
                break
            menu_bar(event, 'raffle')
        except Exception as e:
            pass

    # Close the window
    window.close()


def open_link(link, movie):
    try:
        webbrowser.open(link)
    except:
        sg.popup(
            f"O filme {movie} não possui um link!", title=movie)


def add_movie(gender, film, link):

    for movie in movie_list:
        if movie.lower() == film.lower():
            sg.popup(f'O filme ({film}) ja está na lista!')
            return

    row = len(movie_list) + 2
    # Load the existing workbook
    workbook = openpyxl.load_workbook(file_path)

    # Select the worksheet you want to add data to
    worksheet = workbook.active

    # Add data to the worksheet
    worksheet[f'A{row}'] = gender
    worksheet[f'B{row}'] = film.replace('\n', '').title()
    worksheet[f'C{row}'] = link
    worksheet[f'D{row}'] = 'Não'

    # Save the changes to the workbook
    try:
        workbook.save(file_path)
        att_movies()

        sg.popup(f'O filme ({film.title()}) foi adicionado!')
    except (PermissionError):
        sg.popup('É necessário fechar o arquivo Excel!')
    except Exception as e:
        sg.popup(e)


def edit_movie_window():
    global window
    movies = []
    for movie in movie_list:
        movies.append(movie)
    genders = ['Romance', 'Terror', 'Heróis', 'Animação',
               'Ação/Aventura', 'Suspense']
    layout = [[sg.Menu(menu_layout)],
              [sg.Text('Filme que deseja editar:')],
              [sg.Combo(values=movies, key='-COMBO-', size=(50, 1))],
              [sg.Text('Editar o Gênero do Filme:')],
              [sg.Combo(values=genders, key='-GENDER-', size=(20, 1))],
              [sg.Text('Editar Nome do Filme:'), sg.InputText(key='-FILM-')],
              [sg.Text('Editar Link do Filme:'), sg.InputText(key='-LINK-')],
              [sg.Button('Ok'), sg.Button('Cancelar')]]

    window = sg.Window('Editar Filme', layout, use_ttk_buttons=True)

    while True:
        event, values = window.read()
        try:
            if event == sg.WIN_CLOSED or event == 'Sair':
                break
            elif event == 'Cancelar':
                window.close()
                main_window()
            elif event == 'Ok':
                selected_movie = values['-COMBO-']
                gender_value = values['-GENDER-']
                film_value = values['-FILM-']
                link_value = values['-LINK-']

                if selected_movie != '':
                    if film_value != '' or link_value != '' or gender_value != '':
                        edit_movie(selected_movie, gender_value,
                                   film_value, link_value)
                        window['-COMBO-'].update()

            menu_bar(event, 'edit')
        except:
            pass
    window.close()


def delete_movie_window():
    global window
    global movie_list
    movies = []
    for movie in movie_list:
        movies.append(movie)
    genders = ['Romance', 'Terror', 'Heróis', 'Animação',
               'Ação/Aventura', 'Suspense']
    layout = [[sg.Menu(menu_layout)],
              [sg.Text('Filme que deseja Deletar:')],
              [sg.Combo(values=movies, key='-COMBO-', size=(50, 1))],
              [sg.Button('Ok'), sg.Button('Cancelar')]]

    window = sg.Window('Deletar Filme', layout, use_ttk_buttons=True)

    while True:
        event, values = window.read()
        try:
            if event == sg.WIN_CLOSED or event == 'Sair':
                break
            elif event == 'Cancelar':
                window.close()
                main_window()
            elif event == 'Ok':
                selected_movie = values['-COMBO-']
                delete_movie(selected_movie)
                window['-COMBO-'].update()

            menu_bar(event, 'delete')
        except:
            pass
    window.close()


def edit_movie(selected_movie, gender, film, link):
    workbook = openpyxl.load_workbook(file_path)

    # Select the worksheet you want to add data to
    worksheet = workbook.active

    # Add data to the worksheet

    for index, movie in enumerate(movie_list):
        if movie.lower() == selected_movie.lower():
            index += 2
            if gender != '':
                worksheet[f'A{index}'] = gender
            if film != '':
                worksheet[f'B{index}'] = film.replace('\n', '').title()
            if link != '':
                worksheet[f'C{index}'] = link

    # Save the changes to the workbook
    try:
        workbook.save(file_path)
        att_movies()
        sg.popup(f'O filme ({selected_movie.title()}) foi Editado!')
    except (PermissionError):
        sg.popup('É necessário fechar o arquivo Excel!')
    except Exception as e:
        sg.popup(e)


def nova_planilha(directory_path):
    # Cria um novo arquivo do Excel
    wb = openpyxl.Workbook()

    # Seleciona a planilha ativa
    ws = wb.active
    ws.title = 'List'
    # Define os headers da planilha
    headers = ['Gênero', 'Filme', 'Link', 'Visto']

    # Adiciona os headers na primeira linha da planilha
    ws.append(headers)
    filename = 'lista_filmes.xlsx'
    full_path = os.path.join(directory_path, filename)
    # Salva o arquivo do Excel
    wb.save(full_path)
    return filename, full_path


def delete_movie(selected_movie):
    workbook = openpyxl.load_workbook(file_path)

    for index, movie in enumerate(movie_list):
        if movie.lower() == selected_movie.lower():
            worksheet = workbook['List']
            # Excluir a linha
            worksheet.delete_rows(index+2)

    try:
        workbook.save(file_path)
        att_movies()
        sg.popup(f'O filme ({selected_movie.title()}) foi Deletado!')
    except (PermissionError):
        sg.popup('É necessário fechar o arquivo Excel!')
    except Exception as e:
        sg.popup(e)


def att_movies():
    global movie_list
    global gender_list
    global link_list
    global check_list

    try:
        df = pd.read_excel(file_path, sheet_name='List')
        df_sorted = df.sort_values('Filme')
        df_sorted.to_excel(file_path,
                           index=False, sheet_name='List')

        table = pd.read_excel(file_path, None)

        movie_list = table['List']['Filme']
        gender_list = table['List']['Gênero']
        link_list = table['List']['Link']
        check_list = table['List']['Visto']

    except (PermissionError):
        sg.popup('É necessário fechar o arquivo Excel!')
    except Exception as e:
        sg.popup(e)


def open_file_window():
    global window
    global file_path
    layout = [
        [sg.Menu(menu_layout)],
        [sg.Text('Selecione uma planilha: '), sg.Input(
            key='_FILEBROWSER_'), sg.FileBrowse()],
        [sg.Button('Salvar')]]

    window = sg.Window('Abrir planilha', layout, use_ttk_buttons=True)

    while True:
        event, values = window.read()

        if event == sg.WINDOW_CLOSED:
            break
        elif event == 'Salvar':
            path = values['_FILEBROWSER_']
            if path != '':
                if path.lower().endswith('.xlsx'):
                    # Salva o diretório do arquivo em uma variável
                    directory_path = os.path.dirname(path)
                    filename = os.path.basename(path)
                    full_path = os.path.join(directory_path, filename)
                    file_path = full_path
                    att_movies()
                    sg.popup('Planilha selecionada!')
                else:
                    sg.popup('O Arquivo selecionado não é um arquivo Excel!')
                window.close()
                main_window()
        menu_bar(event, 'open_file')
    window.close()


def new_file_window():
    global window
    global file_path

    layout = [
        [sg.Menu(menu_layout)],
        [sg.Text('Selecione onde quer salvar o arquivo: '), sg.Input(
            key='_FOLDERBROWSER_'), sg.FolderBrowse()],
        [sg.Button('Salvar')]]

    window = sg.Window('Criar planilha', layout, use_ttk_buttons=True)

    while True:
        event, values = window.read()

        if event == sg.WINDOW_CLOSED:
            break
        elif event == 'Salvar':
            folder_path = values['_FOLDERBROWSER_']
            if folder_path != '':
                filename, full_path = nova_planilha(folder_path)
                sg.popup(f'Arquivo {filename} criado!')
                file_path = full_path
                att_movies()
                window.close()
                main_window()
        menu_bar(event, 'new_file')
    window.close()


def menu_bar(event, win):
    global window

    if event == 'Adicionar':
        window.close()
        add_film_window()
    elif event == 'Sorteador':
        window.close()
        main_window()
    elif event == 'Lista':
        window.close()
        list_window()
    elif event == 'Editar':
        window.close()
        edit_movie_window()
    elif event == 'Excluir':
        window.close()
        delete_movie_window()
    elif event == 'Padrão':
        sg.change_look_and_feel('Dark Blue 3')
        detect_window(win)
    elif event == 'Claro':
        sg.change_look_and_feel('LightGreen')
        detect_window(win)
    elif event == 'Escuro':
        sg.change_look_and_feel('DarkAmber')
        detect_window(win)
    elif event == 'Sobre':
        sg.popup('Produzido por: Reinier Soares')
    elif event == 'Abrir':
        window.close()
        open_file_window()
    elif event == 'Criar':
        window.close()
        new_file_window()


def detect_window(win):
    global window
    window.close()

    if win == 'add':
        add_film_window()
    elif win == 'delete':
        delete_movie_window()
    elif win == 'raffle':
        main_window()
    elif win == 'list':
        list_window()
    elif win == 'new_file':
        new_file_window()
    elif win == 'open_file':
        open_file_window()
    elif win == 'edit':
        edit_movie_window()


window = None
link = None
movie = None
movie_list = None
gender_list = None
link_list = None
check_list = None

sep = '---'
menu_layout = [
    ['Filme', ['Sorteador', 'Lista', sep, 'Adicionar', 'Editar',
               'Excluir', sep, 'Sair']],
    ['Planilha', ['Criar', 'Abrir']],
    ['Tema', ['Padrão', 'Claro', 'Escuro']],
    ['Ajuda', ['Sobre']]]

file_path = 'movie_list.xlsx'
try:
    table = pd.read_excel(file_path, None)
except (PermissionError):
    sg.popup('É necessário fechar o arquivo Excel!')
except Exception as e:
    layout = [
        [sg.Text('Arquivo não encontrado, crie ou selecione uma planilha!')],
        [sg.Button('Criar'), sg.Button('Selecionar')]
    ]

    window = sg.Window('Arquivo não encontrado', layout, use_ttk_buttons=True)

    event, values = window.read()

    if event == 'Criar':
        window.close()
        new_file_window()
    elif event == 'Selecionar':
        window.close()
        open_file_window()

att_movies()

main_window()
